import tkinter as tk
import datetime
import random
import string
from tkinter import filedialog, messagebox, simpledialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import threading, time, requests, os, random
from openpyxl import Workbook, load_workbook
from collections import defaultdict

EMOJIS = ["😀", "😃", "😄", "😁", "😆", "😅", "😂", "🤣", "😊", "😇", "🙂", "🙃", "😉", "😌", "😍", "🥰", "😘", "😗", "😙", "😚", "😋", "😛", "😝", "😜", "🤪", "🤨", "🧐", "🤓", "😎", "🤩", "🥳", "😏", "😒", "😞", "😔", "😟", "😕", "🙁", "☹️", "😣", "😖", "😫", "😩", "🥺", "😢", "😭", "😤", "😠", "😡", "🤬", "🤯", "😳", "🥵", "🥶", "😱", "😨", "😰", "😥", "😓", "🤗", "🤔", "🤭", "🤫", "🤥", "😶", "😐", "😑", "😬", "🙄", "😯", "😦", "😧", "😮", "😲", "🥱", "😴", "🤤", "😪", "😵", "🤐", "🥴", "🤢", "🤮", "🤧", "😷", "🤒", "🤕", "🤑", "🤠", "😈", "👿", "👹", "👺", "🤡", "💩", "👻", "💀", "☠️", "👽", "👾", "🤖", "🎃", "😺", "😸", "😹", "😻", "😼", "😽", "🙀", "😿", "😾"]

def process_message(message):
    now = datetime.datetime.utcnow() + datetime.timedelta(hours=7)  
    time_str = now.strftime("%H:%M %d/%m/%Y")
    message = message.replace("{time}", time_str)

    if "{emoji}" in message:
        random_emoji = random.choice(EMOJIS)
        message = message.replace("{emoji}", random_emoji)

    if "{text}" in message:
        length = random.randint(5, 10)
        random_text = ''.join(random.choices(string.ascii_letters + string.digits, k=length))
        message = message.replace("{text}", random_text)

    return message

def get_token_from_cookie(cookie: str):
    url = f"https://adidaphat.site/facebook/tokentocookie?type=EAAAAU&cookie={cookie}"
    try:
        res = requests.get(url, timeout=30)
        data = res.json()
        return data.get("token")
    except Exception:
        return None

def get_post_reactions(post_id, token, proxy=None):
    try:
        proxies = {"http": proxy, "https": proxy} if proxy else None
        url = f"https://graph.facebook.com/{post_id}?fields=reactions.type(LIKE).limit(0).summary(total_count)&access_token={token}"
        res = requests.get(url, proxies=proxies, timeout=30).json()
        if "error" in res:
            return 0
        reactions = res.get("reactions", {})
        return reactions.get("summary", {}).get("total_count", 0)
    except Exception:
        return 0

def get_post_ids(uid, token, limit, proxy=None):
    try:
        proxies = {"http": proxy, "https": proxy} if proxy else None
        url = f"https://graph.facebook.com/{uid}/posts?fields=id,type&limit={limit}&access_token={token}"
        ids = []
        count = 0
        while True:
            res = requests.get(url, proxies=proxies, timeout=30).json()
            if "error" in res:
                break
            data_posts = res.get("data", [])
            for post in data_posts:
                if count < limit:
                    post_type = post.get("type", "")
                    if post_type not in ["video", "reel"]:
                        ids.append(post["id"])
                        count += 1
                else:
                    break
            if count >= limit:
                break
            next_page = res.get("paging", {}).get("next")
            if not next_page:
                break
            url = next_page
        return ids
    except Exception:
        return []

def send_comment(post_id, token, message, image_urls, proxy=None):
    url = f"https://graph.facebook.com/{post_id}/comments"
    data = {"message": message, "access_token": token}
    proxies = {"http": proxy, "https": proxy} if proxy else None
    if image_urls:
        if len(image_urls) == 1:
            data["attachment_url"] = image_urls[0]
    try:
        r = requests.post(url, data=data, timeout=60, proxies=proxies)
        js = r.json()
        cid = None
        if isinstance(js, dict):
            if "id" in js:
                cid = js["id"]
            elif "comment_id" in js:
                cid = js["comment_id"]
        if cid:
            if "_" in str(cid):
                comment_id = str(cid).split("_")[-1]
            else:
                comment_id = str(cid)
            link = f"https://www.facebook.com/{post_id}?comment_id={comment_id}"
            return link
    except Exception:
        pass
    return None

def worker_thread_parallel(selected_posts, tokens_data, messages, images, repeats_per_post, delay, ws_rows, lock, log_fn, stop_event, update_token_log_fn, post_comment_count, token_comment_count, token_limit, min_reactions, daily_limit, daily_comment_count, current_date):
    try:
        today = datetime.datetime.now().date()
        if today != current_date[0]:
            with lock:
                daily_comment_count.clear()
                current_date[0] = today
            log_fn("🔄 Đã reset counter comment hàng ngày")
        
        tasks = []
        for pid in selected_posts:
            tasks.extend([pid] * repeats_per_post)
        random.shuffle(tasks)
        
        available_tokens = []
        for token_info in tokens_data:
            token = token_info['token']
            with lock:
                token_ok = True
                if token_limit > 0 and token_comment_count[token] >= token_limit:
                    token_ok = False
                if daily_limit > 0 and daily_comment_count.get(token, 0) >= daily_limit:
                    token_ok = False
                    log_fn(f"⏹️ Token {token[:20]}... đã đạt giới hạn {daily_limit} comment/ngày")
                
                if token_ok:
                    available_tokens.append(token_info)
        
        if not available_tokens:
            log_fn("⚠️ Tất cả token đã đạt giới hạn comment!")
            return
        
        token_index = 0
        
        for pid in tasks:
            if stop_event.is_set():
                break
                
            with lock:
                if post_comment_count[pid] >= repeats_per_post:
                    continue
            
            attempts = 0
            token_info = None
            while attempts < len(available_tokens):
                token_info = available_tokens[token_index % len(available_tokens)]
                token = token_info['token']
                
                with lock:
                    token_ok = True
                    if token_limit > 0 and token_comment_count[token] >= token_limit:
                        token_ok = False
                    if daily_limit > 0 and daily_comment_count.get(token, 0) >= daily_limit:
                        token_ok = False
                    
                    if token_ok:
                        break
                    else:
                        available_tokens.remove(token_info)
                        if not available_tokens:
                            log_fn("⚠️ Tất cả token đã đạt giới hạn comment!")
                            return
                        continue
                
                token_index += 1
                attempts += 1
            
            if not token_info or attempts >= len(available_tokens):
                log_fn("⚠️ Không tìm thấy token khả dụng!")
                break
            
            token = token_info['token']
            proxy = token_info.get('proxy', '')
            
            if min_reactions > 0:
                try:
                    reaction_count = get_post_reactions(pid, token, proxy)
                    if reaction_count < min_reactions:
                        log_fn(f"⏭️ Bỏ qua {pid} - Chỉ có {reaction_count} reaction (cần {min_reactions})")
                        continue
                    else:
                        log_fn(f"✅ {pid} có {reaction_count} reaction - Đủ điều kiện comment")
                except Exception as e:
                    log_fn(f"⚠️ Không thể kiểm tra reaction của {pid}: {e}")
                    pass
            
            msg = random.choice(messages) if messages else ""
            if msg:
                msg = process_message(msg)
            
            try:
                update_token_log_fn(token, f"💬 Đang comment vào {pid}" + (f" | Proxy: {proxy}" if proxy else ""))
            except Exception:
                pass
            
            link = send_comment(pid, token, msg, images, proxy)
            row = [(token[:20] + '...') if token else '', msg, pid, link or "Thất bại", proxy]
            
            with lock:
                ws_rows.append(row)
                if link:
                    post_comment_count[pid] += 1
                    token_comment_count[token] += 1
                    if daily_limit > 0:
                        daily_comment_count[token] = daily_comment_count.get(token, 0) + 1
                        daily_count = daily_comment_count[token]
                        log_fn(f"📊 Token {token[:20]}... đã cmt {daily_count}/{daily_limit} lượt hôm nay")
            
            if link:
                try:
                    update_token_log_fn(token, f"✅ Đã comment: {link}" + (f" | Proxy: {proxy}" if proxy else ""))
                except Exception:
                    pass
                log_fn(f"{pid} | {token[:20]}... | {msg} | {link}" + (f" | Proxy: {proxy}" if proxy else ""))
                save_link_to_excel(link)
            else:
                try:
                    update_token_log_fn(token, f"❌ Thất bại khi comment vào {pid}" + (f" | Proxy: {proxy}" if proxy else ""))
                except Exception:
                    pass
                log_fn(f"{pid} | {token[:20]}... | {msg} | ❌" + (f" | Proxy: {proxy}" if proxy else ""))
            
            if delay > 0 and not stop_event.is_set():
                try:
                    update_token_log_fn(token, f"Đang chờ {delay} giây..." + (f" | Proxy: {proxy}" if proxy else ""))
                except Exception:
                    pass
                
                end_time = time.time() + delay
                while time.time() < end_time and not stop_event.is_set():
                    remaining = end_time - time.time()
                    if remaining <= 0:
                        break
                    sleep_time = min(0.5, remaining)
                    time.sleep(sleep_time)
            
            if stop_event.is_set():
                try:
                    update_token_log_fn(token, "Đã dừng!")
                except Exception:
                    pass
                break
        
        if not stop_event.is_set():
            for token_info in tokens_data:
                try:
                    update_token_log_fn(token_info['token'], "Hoàn thành!")
                except Exception:
                    pass
    except Exception as e:
        log_fn(f"Lỗi worker: {e}")

def worker_thread_sequential(selected_posts, tokens_data, messages, images, repeats_per_post, delay, ws_rows, lock, log_fn, stop_event, update_token_log_fn, post_comment_count, token_comment_count, token_limit, min_reactions, daily_limit, daily_comment_count, current_date):
    try:
        today = datetime.datetime.now().date()
        if today != current_date[0]:
            with lock:
                daily_comment_count.clear()
                current_date[0] = today
            log_fn("🔄 Đã reset counter comment hàng ngày")
        
        tasks = []
        for pid in selected_posts:
            tasks.extend([pid] * repeats_per_post)
        random.shuffle(tasks)
        
        for token_info in tokens_data:
            if stop_event.is_set():
                break
                    
            token = token_info['token']
            proxy = token_info.get('proxy', '')
            
            with lock:
                token_ok = True
                if token_limit > 0 and token_comment_count[token] >= token_limit:
                    token_ok = False
                if daily_limit > 0 and daily_comment_count.get(token, 0) >= daily_limit:
                    token_ok = False
                
                if not token_ok:
                    update_token_log_fn(token, f"⏹️ Đã đạt giới hạn comment" + (f" | Proxy: {proxy}" if proxy else ""))
                    continue
            
            token_used = False
            for pid in tasks[:]:
                if stop_event.is_set():
                    break
                        
                with lock:
                    if post_comment_count[pid] >= repeats_per_post:
                        tasks = [p for p in tasks if p != pid]
                        continue
                
                with lock:
                    token_ok = True
                    if token_limit > 0 and token_comment_count[token] >= token_limit:
                        token_ok = False
                    if daily_limit > 0 and daily_comment_count.get(token, 0) >= daily_limit:
                        token_ok = False
                    
                    if not token_ok:
                        update_token_log_fn(token, f"⏹️ Đã đạt giới hạn comment" + (f" | Proxy: {proxy}" if proxy else ""))
                        break
                
                if min_reactions > 0:
                    try:
                        reaction_count = get_post_reactions(pid, token, proxy)
                        if reaction_count < min_reactions:
                            log_fn(f"⏭️ Bỏ qua {pid} - Chỉ có {reaction_count} reaction (cần {min_reactions})")
                            if pid in tasks:
                                tasks.remove(pid)
                            continue
                        else:
                            log_fn(f"✅ {pid} có {reaction_count} reaction - Đủ điều kiện comment")
                    except Exception as e:
                        log_fn(f"⚠️ Không thể kiểm tra reaction của {pid}: {e}")
                        pass
                
                msg = random.choice(messages) if messages else ""
                if msg:
                    msg = process_message(msg)
                
                try:
                    update_token_log_fn(token, f"💬 Đang comment vào {pid}" + (f" | Proxy: {proxy}" if proxy else ""))
                except Exception:
                    pass
                
                link = send_comment(pid, token, msg, images, proxy)
                row = [(token[:20] + '...') if token else '', msg, pid, link or "Thất bại", proxy]
                
                with lock:
                    ws_rows.append(row)
                    if link:
                        post_comment_count[pid] += 1
                        token_comment_count[token] += 1
                        token_used = True
                        if daily_limit > 0:
                            daily_comment_count[token] = daily_comment_count.get(token, 0) + 1
                            daily_count = daily_comment_count[token]
                            log_fn(f"📊 Token {token[:20]}... đã cmt {daily_count}/{daily_limit} lượt hôm nay")
                
                if link:
                    try:
                        update_token_log_fn(token, f"✅ Đã comment: {link}" + (f" | Proxy: {proxy}" if proxy else ""))
                    except Exception:
                        pass
                    log_fn(f"{pid} | {token[:20]}... | {msg} | {link}" + (f" | Proxy: {proxy}" if proxy else ""))
                    save_link_to_excel(link)
                else:
                    try:
                        update_token_log_fn(token, f"❌ Thất bại khi comment vào {pid}" + (f" | Proxy: {proxy}" if proxy else ""))
                    except Exception:
                        pass
                    log_fn(f"{pid} | {token[:20]}... | {msg} | ❌" + (f" | Proxy: {proxy}" if proxy else ""))
                
                if delay > 0 and not stop_event.is_set():
                    try:
                        update_token_log_fn(token, f"Đang chờ {delay} giây..." + (f" | Proxy: {proxy}" if proxy else ""))
                    except Exception:
                        pass
                    
                    end_time = time.time() + delay
                    while time.time() < end_time and not stop_event.is_set():
                        remaining = end_time - time.time()
                        if remaining <= 0:
                            break
                        sleep_time = min(0.5, remaining)
                        time.sleep(sleep_time)
                
                if pid in tasks:
                    tasks.remove(pid)
            
            if token_used:
                update_token_log_fn(token, f"✅ Đã hoàn thành lượt comment" + (f" | Proxy: {proxy}" if proxy else ""))
            
            if stop_event.is_set():
                try:
                    update_token_log_fn(token, "Đã dừng!")
                except Exception:
                    pass
                break
        
        if not stop_event.is_set():
            for token_info in tokens_data:
                try:
                    update_token_log_fn(token_info['token'], "Hoàn thành!")
                except Exception:
                    pass
    except Exception as e:
        log_fn(f"Lỗi worker tuần tự: {e}")

def save_link_to_excel(link):
    try:
        filename = "data.xlsx"
        if os.path.exists(filename):
            wb = load_workbook(filename)
            if "Links" in wb.sheetnames:
                ws = wb["Links"]
            else:
                ws = wb.create_sheet("Links")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Links"
        row_num = 1
        while ws.cell(row=row_num, column=1).value is not None:
            row_num += 1
        ws.cell(row=row_num, column=1, value=link)
        wb.save(filename)
    except Exception:
        pass

class TokenManager:
    def __init__(self, parent):
        self.parent = parent
        self.tokens_data = []
        self.selected_tokens = set()
        self.token_logs = {}
        self.proxies = []
        self._is_running = True
        self._threads = []
        self._after_ids = set()
        self._stop_event = threading.Event()
        self.build_ui()

 def _prepare_proxy(self, proxy_str):
    return proxy_str if proxy_str else ""

    def safe_after(self, delay_ms, callback, *args):
        if not self._is_running or self._stop_event.is_set():
            return None
        if not (hasattr(self.parent, 'winfo_exists') and self.parent.winfo_exists()):
            return None
        after_id_holder = {}
        def wrapper():
            try:
                callback(*args)
            finally:
                try:
                    self._after_ids.discard(after_id_holder.get('id'))
                except:
                    pass
        try:
            after_id = self.parent.after(delay_ms, wrapper)
            after_id_holder['id'] = after_id
            self._after_ids.add(after_id)
            return after_id
        except Exception:
            return None

    def cancel_all_afters(self):
        try:
            for aid in list(self._after_ids):
                try:
                    self.parent.after_cancel(aid)
                except Exception:
                    pass
            self._after_ids.clear()
        except Exception:
            pass

    def build_ui(self):
        main_frame = tb.Frame(self.parent)
        main_frame.pack(fill=BOTH, expand=YES, padx=15, pady=15)
        header_frame = tb.Frame(main_frame)
        header_frame.pack(fill=X, pady=(0, 15))
        tb.Label(header_frame, text="QUẢN LÝ TOKEN", font=('Helvetica', 16, 'bold'), bootstyle="primary").pack(side=LEFT)
        control_frame = tb.LabelFrame(main_frame, text="QUẢN LÝ TOKEN", padding=10)
        control_frame.pack(fill=X, pady=(0, 15))
        token_btn_frame = tb.Frame(control_frame)
        token_btn_frame.pack(fill=X, pady=(0, 8))
        tb.Button(token_btn_frame, text="Thêm Token", command=self.add_token, bootstyle="secondary", width=15).pack(side=LEFT, padx=(0, 8))
        tb.Button(token_btn_frame, text="Thêm Cookie", command=self.add_token_from_cookie, bootstyle="warning-outline", width=15).pack(side=LEFT, padx=(0, 8))
        tb.Button(token_btn_frame, text="Thêm Token Từ File", command=self.load_tokens_from_file_dialog, bootstyle="info-outline", width=19).pack(side=LEFT, padx=(0, 8))
        tb.Button(token_btn_frame, text="Xóa Token", command=self.delete_tokens, bootstyle="danger-outline", width=15).pack(side=LEFT, padx=(0, 8))
        tb.Button(token_btn_frame, text="Chọn Tất Cả", command=self.select_all_tokens, bootstyle="success-outline", width=15).pack(side=LEFT, padx=(0, 8))
        tb.Button(token_btn_frame, text="Bỏ Chọn Tất Cả", command=self.deselect_all_tokens, bootstyle="secondary-outline", width=15).pack(side=LEFT)
        proxy_btn_frame = tb.Frame(control_frame)
        proxy_btn_frame.pack(fill=X)
        tb.Button(proxy_btn_frame, text="Thêm Proxy", command=self.add_proxy, bootstyle="secondary", width=15).pack(side=LEFT, padx=(0, 8))
        tb.Button(proxy_btn_frame, text="Thêm Proxy Từ File", command=self.load_proxies_from_file, bootstyle="info-outline", width=19).pack(side=LEFT, padx=(0, 8))
        tb.Button(proxy_btn_frame, text="Tự Gán Proxy", command=self.auto_assign_proxies, bootstyle="warning-outline", width=15).pack(side=LEFT, padx=(0, 8))
        tb.Button(proxy_btn_frame, text="Xóa Proxy", command=self.clear_all_proxies, bootstyle="danger-outline", width=15).pack(side=LEFT)
        tree_container = tb.LabelFrame(main_frame, text="DANH SÁCH TOKEN", padding=10)
        tree_container.pack(fill=BOTH, expand=YES)
        columns = ("select", "uid", "name", "token", "proxy", "status", "log")
        self.tree = tb.Treeview(tree_container, columns=columns, show="headings", height=12)
        self.tree.heading("select", text="CHỌN", anchor="center")
        self.tree.heading("uid", text="UID", anchor="center")
        self.tree.heading("name", text="TÊN", anchor="center")
        self.tree.heading("token", text="TOKEN", anchor="center")
        self.tree.heading("proxy", text="PROXY", anchor="center")
        self.tree.heading("status", text="TRẠNG THÁI", anchor="center")
        self.tree.heading("log", text="NHẬT KÝ", anchor="center")
        self.tree.column("select", width=50, anchor="center", stretch=False)
        self.tree.column("uid", width=100, anchor="center", stretch=False)
        self.tree.column("name", width=150, anchor="center", stretch=False)
        self.tree.column("token", width=250, anchor="center", stretch=False)
        self.tree.column("proxy", width=180, anchor="center", stretch=False)
        self.tree.column("status", width=100, anchor="center", stretch=False)
        self.tree.column("log", width=200, anchor="center", stretch=True)
        v_scrollbar = tb.Scrollbar(tree_container, orient=VERTICAL, command=self.tree.yview)
        h_scrollbar = tb.Scrollbar(tree_container, orient=HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        self.tree.bind('<Button-1>', self.on_tree_click)
        self.tree.bind('<Button-3>', self.show_context_menu)

    def cleanup(self):
        self._is_running = False
        try:
            self._stop_event.set()
        except:
            pass
        try:
            self.cancel_all_afters()
        except:
            pass
        for t in list(self._threads):
            try:
                t.join(timeout=1.0)
            except Exception:
                pass
        self._threads.clear()

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            context_menu = tk.Menu(self.parent, tearoff=0)
            context_menu.add_command(label="📋 Copy Token", command=lambda: self.copy_selected_field("token"))
            context_menu.add_command(label="🔢 Copy UID", command=lambda: self.copy_selected_field("uid"))
            context_menu.add_command(label="👤 Copy Name", command=lambda: self.copy_selected_field("name"))
            context_menu.add_command(label="🌐 Copy Proxy", command=lambda: self.copy_selected_field("proxy"))
            context_menu.add_separator()
            context_menu.add_command(label="➕ Add Proxy", command=self.add_proxy_to_selected)
            context_menu.add_command(label="➖ Remove Proxy", command=self.remove_proxy_from_selected)
            context_menu.add_separator()
            context_menu.add_command(label="📄 Copy All Info", command=self.copy_all_info)
            try:
                context_menu.post(event.x_root, event.y_root)
            except Exception:
                pass

    def add_proxy_to_selected(self):
        selected_items = self.tree.selection()
        if not selected_items:
            try:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một token!")
            except:
                pass
            return
        proxy = simpledialog.askstring("Thêm Proxy", "Nhập proxy IPv6:\n\nĐịnh dạng:\n- host:port:username:password\n- host:port\n- http://user:pass@host:port")
        if proxy and proxy.strip():
            formatted_proxy = self._prepare_proxy(proxy.strip())
            for item in selected_items:
                self.update_token_proxy(item, formatted_proxy)

    def remove_proxy_from_selected(self):
        selected_items = self.tree.selection()
        if not selected_items:
            try:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một token!")
            except:
                pass
            return
        for item in selected_items:
            self.update_token_proxy(item, "")

    def update_token_proxy(self, item, proxy):
        values = list(self.tree.item(item)['values'])
        if len(values) >= 5:
            values[4] = proxy
            try:
                self.tree.item(item, values=values)
            except Exception:
                pass

    def copy_selected_field(self, field):
        selected_items = self.tree.selection()
        if selected_items:
            item = selected_items[0]
            values = self.tree.item(item)['values']
            if field == "token" and len(values) > 3:
                self.copy_to_clipboard(values[3])
            elif field == "uid" and len(values) > 1:
                self.copy_to_clipboard(values[1])
            elif field == "name" and len(values) > 2:
                self.copy_to_clipboard(values[2])
            elif field == "proxy" and len(values) > 4:
                self.copy_to_clipboard(values[4])

    def copy_all_info(self):
        selected_items = self.tree.selection()
        if selected_items:
            item = selected_items[0]
            values = self.tree.item(item)['values']
            if len(values) >= 5:
                info = f"UID: {values[1]}\nName: {values[2]}\nToken: {values[3]}\nProxy: {values[4]}\nStatus: {values[5]}"
                self.copy_to_clipboard(info)

    def copy_to_clipboard(self, text):
        try:
            self.parent.clipboard_clear()
            self.parent.clipboard_append(str(text))
        except Exception:
            pass

    def on_tree_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            item = self.tree.identify_row(event.y)
            if column == "#1":
                self.toggle_selection(item)

    def toggle_selection(self, item):
        if item:
            values = list(self.tree.item(item)['values'])
            if values:
                token_index = self.tree.index(item)
                if token_index in self.selected_tokens:
                    self.selected_tokens.remove(token_index)
                    values[0] = "☐"
                else:
                    self.selected_tokens.add(token_index)
                    values[0] = "☑"
                try:
                    self.tree.item(item, values=values)
                except Exception:
                    pass

    def add_token(self):
        token = simpledialog.askstring("Thêm Token", "Nhập token Facebook:")
        if token and token.strip():
            if self.token_exists(token.strip()):
                try:
                    messagebox.showwarning("Token đã tồn tại", "Token này đã có trong danh sách!")
                except:
                    pass
                return
            self.check_and_add_token(token.strip())

    def add_token_from_cookie(self):
        cookie = simpledialog.askstring("Thêm Từ Cookie", "Nhập cookie Facebook:")
        if cookie and cookie.strip():
            def convert_cookie():
                try:
                    token = get_token_from_cookie(cookie.strip())
                    if token:
                        self.safe_after(0, lambda: self.process_converted_token(token, cookie))
                    else:
                        self.safe_after(0, lambda: messagebox.showerror("Lỗi", "Không thể chuyển đổi cookie thành token!"))
                except Exception:
                    self.safe_after(0, lambda: messagebox.showerror("Lỗi", "Lỗi khi chuyển đổi"))
            self.safe_after(0, lambda: self.add_token_to_table("Đang chuyển đổi...", "Đang chuyển đổi...", "Đang chuyển đổi...", "Converting", "", "Đang chuyển cookie thành token..."))
            t = threading.Thread(target=convert_cookie, daemon=True)
            self._threads.append(t)
            t.start()

    def process_converted_token(self, token, original_cookie):
        try:
            for item in self.tree.get_children():
                item_values = self.tree.item(item)['values']
                if len(item_values) > 3 and item_values[3] == "Đang chuyển đổi..." and item_values[5] == "Converting":
                    try:
                        self.tree.delete(item)
                    except Exception:
                        pass
                    break
        except Exception:
            pass
        if self.token_exists(token):
            try:
                messagebox.showwarning("Token đã tồn tại", "Token này đã có trong danh sách!")
            except:
                pass
            return
        self.check_and_add_token(token)

    def token_exists(self, token):
        try:
            for item in self.tree.get_children():
                item_values = self.tree.item(item)['values']
                if len(item_values) > 3 and item_values[3] == token:
                    return True
        except Exception:
            pass
        return False

    def check_and_add_token(self, token):
        def check_token():
            try:
                if self._stop_event.is_set():
                    return
                url = f"https://graph.facebook.com/me?fields=id,name&access_token={token}"
                response = requests.get(url, timeout=30)
                if self._stop_event.is_set():
                    return
                data = response.json()
                if isinstance(data, dict) and 'id' in data and 'name' in data:
                    uid = data['id']
                    name = data['name']
                    status = "Live"
                    log = "Token hoạt động"
                else:
                    uid = "N/A"
                    name = "N/A"
                    status = "Die"
                    log = data.get('error', {}).get('message', 'Lỗi không xác định') if isinstance(data, dict) else 'Lỗi không xác định'
                if self._stop_event.is_set():
                    return
                self.safe_after(0, lambda uid=uid, name=name, status=status, log=log: self.add_token_to_table(token, uid, name, status, "", log))
            except Exception as e:
                error_msg = str(e)
                if not self._stop_event.is_set():
                    self.safe_after(0, lambda error_msg=error_msg: self.add_token_to_table(token, "N/A", "N/A", "Die", "", f"Lỗi: {error_msg}"))
        self.safe_after(0, lambda: self.add_token_to_table(token, "Đang kiểm tra...", "Đang kiểm tra...", "Checking", "", "Đang xác thực token..."))
        t = threading.Thread(target=check_token, daemon=True)
        self._threads.append(t)
        t.start()
    
    def select_all_tokens(self):
        for item in self.tree.get_children():
            values = list(self.tree.item(item)['values'])
            token_index = self.tree.index(item)
            self.selected_tokens.add(token_index)
            values[0] = "☑"
            try:
                self.tree.item(item, values=values)
            except Exception:
                pass

    def deselect_all_tokens(self):
        for item in self.tree.get_children():
            values = list(self.tree.item(item)['values'])
            token_index = self.tree.index(item)
            if token_index in self.selected_tokens:
                self.selected_tokens.remove(token_index)
            values[0] = "☐"
            try:
                self.tree.item(item, values=values)
            except Exception:
                pass

    def add_token_to_table(self, token, uid, name, status, proxy, log):
        if not self._is_running:
            return
        def do_add():
            try:
                existing_item = None
                for item in self.tree.get_children():
                    item_values = self.tree.item(item)['values']
                    if len(item_values) > 3 and item_values[3] == token and (item_values[5] in ("Checking", "Converting")):
                        existing_item = item
                        break
                values = ("☐", uid, name, token, proxy, status, log)
                if existing_item:
                    try:
                        self.tree.item(existing_item, values=values)
                    except Exception:
                        pass
                else:
                    try:
                        self.tree.insert("", tk.END, values=values)
                    except Exception:
                        pass
                try:
                    self.token_logs[token] = log
                except Exception:
                    pass
            except Exception:
                pass
        self.safe_after(0, do_add)

    def update_token_log(self, token, log_message):
        try:
            self.token_logs[token] = log_message
        except Exception:
            pass
        def do_update():
            try:
                if not hasattr(self, 'tree'):
                    return
                for item in self.tree.get_children():
                    item_values = self.tree.item(item)['values']
                    if len(item_values) > 3 and item_values[3] == token:
                        new_values = list(item_values)
                        if len(new_values) < 7:
                            new_values += [''] * (7 - len(new_values))
                        new_values[6] = log_message
                        try:
                            self.tree.item(item, values=new_values)
                        except Exception:
                            pass
                        break
            except Exception:
                pass
        self.safe_after(0, do_update)

    def add_proxy(self):
        proxy = simpledialog.askstring("Thêm Proxy", "Nhập proxy IPv6:\n\nĐịnh dạng:\n- host:port:username:password\n- host:port\n- http://user:pass@host:port")
        if proxy and proxy.strip():
            formatted_proxy = self._prepare_proxy(proxy.strip())
            self.proxies.append(formatted_proxy)
            try:
                messagebox.showinfo("Thành công", f"Đã thêm proxy: {formatted_proxy}")
            except Exception:
                pass

    def load_proxies_from_file(self):
        filename = filedialog.askopenfilename(title="Chọn file proxy", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    raw_proxies = [line.strip() for line in f if line.strip()]
                formatted_proxies = [self._prepare_proxy(proxy) for proxy in raw_proxies]
                self.proxies.extend(formatted_proxies)
                try:
                    messagebox.showinfo("Thành công", f"Đã load {len(formatted_proxies)} proxy từ file")
                except Exception:
                    pass
            except Exception:
                pass

    def auto_assign_proxies(self):
        if not self.proxies:
            try:
                messagebox.showwarning("Cảnh báo", "Chưa có proxy nào trong danh sách!")
            except:
                pass
            return
        live_tokens = []
        for item in self.tree.get_children():
            item_values = self.tree.item(item)['values']
            if len(item_values) > 5 and item_values[5] == "Live":
                live_tokens.append(item)
        if not live_tokens:
            try:
                messagebox.showwarning("Cảnh báo", "Không có token Live nào để gán proxy!")
            except:
                pass
            return
        random.shuffle(self.proxies)
        assigned_count = 0
        for i, item in enumerate(live_tokens):
            proxy = self.proxies[i % len(self.proxies)]
            self.update_token_proxy(item, proxy)
            assigned_count += 1
        try:
            messagebox.showinfo("Thành công", f"Đã gán proxy cho {assigned_count} token")
        except:
            pass

    def clear_all_proxies(self):
        try:
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa tất cả proxy khỏi token?"):
                for item in self.tree.get_children():
                    self.update_token_proxy(item, "")
                messagebox.showinfo("Thành công", "Đã xóa tất cả proxy")
        except Exception:
            pass

    def load_tokens_from_file_dialog(self):
        filename = filedialog.askopenfilename(title="Chọn file token", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if filename:
            self.load_tokens_from_file(filename)

    def load_tokens_from_file(self, filename=None):
        if not filename:
            filename = "tokens_backup.txt"
        try:
            if os.path.exists(filename):
                with open(filename, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f if line.strip()]
                added_tokens = set()
                for line in lines:
                    parts = line.split('|')
                    token = parts[0].strip()
                    proxy = parts[1].strip() if len(parts) > 1 else ""
                    if token not in added_tokens and not self.token_exists(token):
                        self.check_and_add_token_with_proxy(token, proxy)
                        added_tokens.add(token)
                try:
                    messagebox.showinfo("Thành công", f"Đã load {len(added_tokens)} token từ file")
                except:
                    pass
            else:
                with open("tokens_backup.txt", 'w', encoding='utf-8') as f:
                    pass
        except Exception:
            pass

    def check_and_add_token_with_proxy(self, token, proxy):
        def check_token():
            try:
                if self._stop_event.is_set():
                    return
                url = f"https://graph.facebook.com/me?fields=id,name&access_token={token}"
                response = requests.get(url, timeout=10)
                if self._stop_event.is_set():
                    return
                data = response.json()
                if 'id' in data and 'name' in data:
                    uid = data['id']
                    name = data['name']
                    status = "Live"
                    log = "Token hoạt động"
                else:
                    uid = "N/A"
                    name = "N/A"
                    status = "Die"
                    log = data.get('error', {}).get('message', 'Lỗi không xác định')
                if self._stop_event.is_set():
                    return
                self.safe_after(0, lambda uid=uid, name=name, status=status, log=log: self.add_token_to_table(token, uid, name, status, proxy, log))
            except Exception:
                if not self._stop_event.is_set():
                    self.safe_after(0, lambda: self.add_token_to_table(token, "N/A", "N/A", "Die", proxy, "Lỗi"))
        self.safe_after(0, lambda: self.add_token_to_table(token, "Đang kiểm tra...", "Đang kiểm tra...", "Checking", proxy, "Đang xác thực token..."))
        t = threading.Thread(target=check_token, daemon=True)
        self._threads.append(t)
        t.start()

    def delete_tokens(self):
        if not self.selected_tokens:
            try:
                messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một token để xóa")
            except:
                pass
            return
        try:
            if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa {len(self.selected_tokens)} token?"):
                selected_indices = sorted(self.selected_tokens, reverse=True)
                all_items = list(self.tree.get_children())
                for index in selected_indices:
                    if index < len(all_items):
                        item_values = self.tree.item(all_items[index])['values']
                        if len(item_values) > 3:
                            token = item_values[3]
                            if token in self.token_logs:
                                del self.token_logs[token]
                        try:
                            self.tree.delete(all_items[index])
                        except Exception:
                            pass
                self.selected_tokens.clear()
        except Exception:
            pass

    def save_tokens_to_file(self):
        try:
            tokens_data = []
            for item in self.tree.get_children():
                item_values = self.tree.item(item)['values']
                if len(item_values) > 3:
                    token = item_values[3]
                    proxy = item_values[4] if len(item_values) > 4 else ""
                    tokens_data.append(f"{token}|{proxy}")
            with open("tokens_backup.txt", 'w', encoding='utf-8') as f:
                for data in tokens_data:
                    f.write(data + '\n')
        except Exception:
            pass

    def get_selected_tokens_data(self):
        tokens_data = []
        all_items = list(self.tree.get_children())
        for index in self.selected_tokens:
            if index < len(all_items):
                item_values = self.tree.item(all_items[index])['values']
                if len(item_values) > 3 and item_values[5] == "Live":
                    token_data = {'token': item_values[3], 'proxy': item_values[4] if len(item_values) > 4 else ""}
                    tokens_data.append(token_data)
        return tokens_data

    def get_all_tokens_data(self):
        tokens_data = []
        for item in self.tree.get_children():
            item_values = self.tree.item(item)['values']
            if len(item_values) > 3 and item_values[5] == "Live":
                token_data = {'token': item_values[3], 'proxy': item_values[4] if len(item_values) > 4 else ""}
                tokens_data.append(token_data)
        return tokens_data

class FBCommentGUI:
    def __init__(self, root):
        self.root = root
        self.style = tb.Style(theme="superhero")
        self.message_var = tk.StringVar()
        self.repeats_var = tk.IntVar(value=1)
        self.delay_var = tk.DoubleVar(value=1.0)
        self.threads_var = tk.IntVar(value=2)
        self.limit_var = tk.IntVar(value=100)
        self.token_limit_var = tk.IntVar(value=10)
        self.min_reactions_var = tk.IntVar(value=0)
        self.daily_limit_var = tk.IntVar(value=0)
        self.run_mode = tk.StringVar(value="parallel")
        self.uid_frames = {}
        self.stop_event = threading.Event()
        self.ws_rows = []
        self.lock = threading.Lock()
        self.wb = None
        self.ws = None
        self.token_manager = None
        self.uid_source_text = None
        self._is_running = True
        self._threads = []
        self._after_ids = set()
        self.build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def safe_after(self, delay_ms, callback, *args):
        if not self._is_running:
            return None
        if not (hasattr(self.root, 'winfo_exists') and self.root.winfo_exists()):
            return None
        after_id_holder = {}
        def wrapper():
            try:
                callback(*args)
            finally:
                try:
                    self._after_ids.discard(after_id_holder.get('id'))
                except:
                    pass
        try:
            after_id = self.root.after(delay_ms, wrapper)
            after_id_holder['id'] = after_id
            self._after_ids.add(after_id)
            return after_id
        except Exception:
            return None

    def cancel_all_afters(self):
        try:
            for aid in list(self._after_ids):
                try:
                    self.root.after_cancel(aid)
                except Exception:
                    pass
            self._after_ids.clear()
        except Exception:
            pass

    def on_closing(self):
        self._is_running = False
        self.stop_event.set()
        if self.token_manager:
            try:
                self.token_manager.cleanup()
            except Exception:
                pass
        try:
            self.cancel_all_afters()
        except Exception:
            pass
        for t in list(self._threads):
            try:
                t.join(timeout=1.0)
            except Exception:
                pass
        self._threads.clear()
        try:
            self.root.destroy()
        except Exception:
            pass

    def build_ui(self):
        root = self.root
        root.title("Facebook Truong An")
        try:
            root.state('zoomed')
        except Exception:
            pass
        notebook = tb.Notebook(root, bootstyle="primary")
        notebook.pack(fill=BOTH, expand=YES, padx=15, pady=15)
        comment_tab = tb.Frame(notebook, padding=10)
        notebook.add(comment_tab, text="Chức Năng")
        token_tab = tb.Frame(notebook, padding=10)
        notebook.add(token_tab, text="Quản Lý Token")
        log_tab = tb.Frame(notebook, padding=10)
        notebook.add(log_tab, text="Nhật Ký")
        self.token_manager = TokenManager(token_tab)
        self.build_comment_interface(comment_tab)
        self.build_log_interface(log_tab)

    def build_comment_interface(self, parent):
        main_container = tb.Frame(parent)
        main_container.pack(fill=BOTH, expand=YES)
        paned = tb.PanedWindow(main_container, orient=HORIZONTAL)
        paned.pack(fill=BOTH, expand=YES)
        left_panel = tb.Frame(paned, padding=10)
        right_panel = tb.Frame(paned, width=400, padding=10)
        paned.add(left_panel, weight=3)
        paned.add(right_panel, weight=1)
        uid_frame = tb.LabelFrame(left_panel, text="NHẬP UID HOẶC LINK", padding=15)
        uid_frame.pack(fill=X, pady=(0, 10))
        tb.Label(uid_frame, text="Mỗi dòng 1 UID hoặc link Facebook:", font=('Helvetica', 10, 'bold')).pack(anchor=W, pady=(0, 8))
        self.uid_source_text = tb.Text(uid_frame, height=4, font=('Consolas', 10))
        self.uid_source_text.pack(fill=X, pady=(0, 10))
        uid_control_frame = tb.Frame(uid_frame)
        uid_control_frame.pack(fill=X)
        tb.Button(uid_control_frame, text="Tải ID Bài Viết", command=self.on_fetch, bootstyle="success-outline", width=15).pack(side=LEFT, padx=(0, 10))
        tb.Label(uid_control_frame, text="Số lượng bài viết:").pack(side=LEFT, padx=(10, 5))
        tb.Entry(uid_control_frame, textvariable=self.limit_var, width=8, font=('Helvetica', 10)).pack(side=LEFT, padx=(0, 15))
        tb.Label(uid_control_frame, text="Reaction tối thiểu:").pack(side=LEFT, padx=(1, 5))
        tb.Entry(uid_control_frame, textvariable=self.min_reactions_var, width=8, font=('Helvetica', 10)).pack(side=LEFT, padx=(0, 15))
        tb.Button(uid_control_frame, text="Chọn Tất Cả", command=self.select_all,
          bootstyle="outline", width=17).pack(side=LEFT, padx=(0, 10))
        tb.Button(uid_control_frame, text="Bỏ Chọn Tất Cả", command=self.deselect_all,
                bootstyle="danger-outline", width=17).pack(side=LEFT)
        posts_frame = tb.LabelFrame(left_panel, text="DANH SÁCH ID BÀI VIẾT", padding=15)
        posts_frame.pack(fill=BOTH, expand=YES, pady=(0, 10))
        posts_container = tb.Frame(posts_frame)
        posts_container.pack(fill=BOTH, expand=YES)
        canvas = tk.Canvas(posts_container, bg=self.style.colors.bg)
        scrollbar = tb.Scrollbar(posts_container, orient="vertical", command=canvas.yview)
        self.posts_inner = tb.Frame(canvas)
        self.posts_inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.posts_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=YES)
        scrollbar.pack(side=RIGHT, fill=Y)
        config_frame = tb.LabelFrame(right_panel, text="CẤU HÌNH", padding=15)
        config_frame.pack(fill=X, pady=(0, 10))
        
        placeholder_frame = tb.Frame(config_frame)
        placeholder_frame.pack(fill=X, pady=(0, 10))
        tb.Label(placeholder_frame, text="Placeholders:", font=('Helvetica', 10, 'bold')).pack(anchor=W, pady=(0, 5))
        btn_frame = tb.Frame(placeholder_frame)
        btn_frame.pack(fill=X)
        tb.Button(btn_frame, text="{time}", command=lambda: self.insert_placeholder("{time}"), 
                 bootstyle="outline", width=8).pack(side=LEFT, padx=(0, 5))
        tb.Button(btn_frame, text="{emoji}", command=lambda: self.insert_placeholder("{emoji}"), 
                 bootstyle="outline", width=8).pack(side=LEFT, padx=(0, 5))
        tb.Button(btn_frame, text="{text}", command=lambda: self.insert_placeholder("{text}"), 
                 bootstyle="outline", width=8).pack(side=LEFT)
        
        tb.Label(config_frame, text="Nội dung comment (phân cách bằng |):", font=('Helvetica', 10, 'bold')).pack(anchor=W, pady=(0, 5))
        tb.Entry(config_frame, textvariable=self.message_var, font=('Helvetica', 10)).pack(fill=X, pady=(0, 10))
        tb.Label(config_frame, text="Link ảnh đính kèm (mỗi dòng 1 link):", font=('Helvetica', 10, 'bold')).pack(anchor=W, pady=(0, 5))
        self.images_text = tb.Text(config_frame, height=3, font=('Consolas', 9))
        self.images_text.pack(fill=X, pady=(0, 10))
        params_frame = tb.LabelFrame(config_frame, text="Cài Đặt", padding=10)
        params_frame.pack(fill=X)
        
        tb.Label(params_frame, text="Số lần cmt mỗi bài:").grid(row=0, column=0, sticky=W, pady=5)
        tb.Entry(params_frame, textvariable=self.repeats_var, width=8).grid(row=0, column=1, pady=5, padx=(5, 15))
        
        tb.Label(params_frame, text="Delay (giây):").grid(row=0, column=2, sticky=W, pady=5)
        tb.Entry(params_frame, textvariable=self.delay_var, width=8).grid(row=0, column=3, pady=5, padx=5)
        
        tb.Label(params_frame, text="Số luồng:").grid(row=1, column=0, sticky=W, pady=5)
        tb.Entry(params_frame, textvariable=self.threads_var, width=8).grid(row=1, column=1, pady=5, padx=(5, 15))
        
        tb.Label(params_frame, text="Giới hạn cmt/token:").grid(row=1, column=2, sticky=W, pady=5)
        tb.Entry(params_frame, textvariable=self.token_limit_var, width=8).grid(row=1, column=3, pady=5, padx=5)
        
        tb.Label(params_frame, text="Reaction tối thiểu:").grid(row=2, column=0, sticky=W, pady=5)
        tb.Entry(params_frame, textvariable=self.min_reactions_var, width=8).grid(row=2, column=1, pady=5, padx=(5, 15))
        
        tb.Label(params_frame, text="Giới hạn cmt/ngày:").grid(row=2, column=2, sticky=W, pady=5)
        tb.Entry(params_frame, textvariable=self.daily_limit_var, width=8).grid(row=2, column=3, pady=5, padx=5)
        
        mode_frame = tb.Frame(params_frame)
        mode_frame.grid(row=3, column=0, columnspan=4, sticky=W, pady=5)
        tb.Label(mode_frame, text="Chế độ:").pack(side=LEFT, padx=(0, 10))
        tb.Radiobutton(mode_frame, text="Đồng loạt", variable=self.run_mode, 
                      value="parallel", bootstyle="secondary").pack(side=LEFT, padx=(0, 10))
        tb.Radiobutton(mode_frame, text="Tuần tự", variable=self.run_mode, 
                      value="sequential", bootstyle="secondary").pack(side=LEFT)
        
        control_frame = tb.LabelFrame(right_panel, text="ĐIỀU KHIỂN", padding=15)
        control_frame.pack(fill=X, pady=(0, 10))
        btn_frame = tb.Frame(control_frame)
        btn_frame.pack(fill=X)
        self.start_btn = tb.Button(btn_frame, text="Chạy", command=self.on_start, bootstyle="success-outline", width=12)
        self.start_btn.pack(side=LEFT, padx=(0, 10))
        self.stop_btn = tb.Button(btn_frame, text="Dừng", command=self.on_stop, bootstyle="danger-outline", width=12, state=DISABLED)
        self.stop_btn.pack(side=LEFT, padx=(0, 10))
        tb.Button(btn_frame, text="Lưu excel", command=self.on_save_excel, bootstyle="warning-outline", width=12).pack(side=LEFT)
        status_frame = tb.LabelFrame(right_panel, text="TRẠNG THÁI", padding=15)
        status_frame.pack(fill=BOTH, expand=YES)
        self.status_label = tb.Label(status_frame, text="🟢 SẴN SÀNG", font=('Helvetica', 11, 'bold'), bootstyle="success")
        self.status_label.pack(anchor=W)
        progress_frame = tb.Frame(status_frame)
        progress_frame.pack(fill=X, pady=(10, 0))
        self.stats_label = tb.Label(status_frame, text="Chúc bạn sử dụng tool vui vẻ!", font=('Helvetica', 9), justify=LEFT)
        self.stats_label.pack(anchor=W, pady=(5, 0))

    def build_log_interface(self, parent):
        main_frame = tb.Frame(parent)
        main_frame.pack(fill=BOTH, expand=YES, padx=10, pady=10)
        header_frame = tb.Frame(main_frame)
        header_frame.pack(fill=X, pady=(0, 15))
        tb.Label(header_frame, text="NHẬT KÝ HOẠT ĐỘNG", font=('Helvetica', 16, 'bold'), bootstyle="primary").pack(side=LEFT)
        control_frame = tb.Frame(header_frame)
        control_frame.pack(side=RIGHT)
        tb.Button(control_frame, text="Sao Chép", command=self.copy_log, bootstyle="info-outline", width=12).pack(side=LEFT, padx=(0, 5))
        tb.Button(control_frame, text="Xóa Log", command=self.clear_log, bootstyle="danger-outline", width=12).pack(side=LEFT, padx=(0, 5))
        tb.Button(control_frame, text="Lưu Log", command=self.save_log, bootstyle="success-outline", width=12).pack(side=LEFT)
        log_container = tb.LabelFrame(main_frame, text="NHẬT KÝ CHI TIẾT", padding=10)
        log_container.pack(fill=BOTH, expand=YES)
        self.log_text = tb.Text(log_container, height=25, font=('Consolas', 10), wrap=tk.WORD)
        v_scrollbar = tb.Scrollbar(log_container, orient=VERTICAL, command=self.log_text.yview)
        h_scrollbar = tb.Scrollbar(log_container, orient=HORIZONTAL, command=self.log_text.xview)
        self.log_text.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        log_container.grid_rowconfigure(0, weight=1)
        log_container.grid_columnconfigure(0, weight=1)
        stats_frame = tb.LabelFrame(main_frame, text="THỐNG KÊ HOẠT ĐỘNG", padding=10)
        stats_frame.pack(fill=X, pady=(10, 0))
        stats_container = tb.Frame(stats_frame)
        stats_container.pack(fill=X)
        self.stats_info_label = tb.Label(stats_container, text="🟢 Chưa có thống kê", font=('Helvetica', 10), justify=LEFT)
        self.stats_info_label.pack(anchor=W)
        self.log_text.configure(state=DISABLED)
    
    def save_links_to_excel(self, filename):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Links"
            ws['A1'] = "Danh sách link"
            
            row_num = 2
            for row in self.ws_rows:
                if len(row) > 3:
                    link = row[3]
                    if link and link != "Thất bại":
                        ws.cell(row=row_num, column=1, value=link)
                        row_num += 1
            
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Lỗi khi lưu file Excel: {e}")
            return False

    def copy_log(self):
        try:
            log_content = self.log_text.get('1.0', tk.END)
            if log_content.strip():
                self.root.clipboard_clear()
                self.root.clipboard_append(log_content)
                messagebox.showinfo("Thành công", "Đã sao chép log vào clipboard!")
            else:
                messagebox.showwarning("Cảnh báo", "Không có nội dung log để sao chép!")
        except Exception:
            pass

    def clear_log(self):
        try:
            if messagebox.askyesno("Xác nhận", "Bạn có chắc muốn xóa toàn bộ nhật ký?"):
                self.log_text.configure(state=tk.NORMAL)
                self.log_text.delete('1.0', tk.END)
                self.log_text.configure(state=tk.DISABLED)
                self.stats_info_label.configure(text="🟢 Log đã được xóa")
        except Exception:
            pass

    def save_log(self):
        try:
            log_content = self.log_text.get('1.0', tk.END)
            if not log_content.strip():
                try:
                    messagebox.showwarning("Cảnh báo", "Không có nội dung log để lưu!")
                except:
                    pass
                return
            filename = filedialog.asksaveasfilename(defaultextension='.txt', filetypes=[('Text files', '*.txt'), ('All files', '*.*')], initialfile=f'comment_log_{int(time.time())}.txt')
            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(log_content)
                try:
                    messagebox.showinfo("Thành công", f"Đã lưu log vào file: {filename}")
                except:
                    pass
        except Exception:
            pass

    def log(self, msg):
        if not self._is_running:
            return
        ts = time.strftime('%H:%M:%S')
        formatted_msg = f"[{ts}] {msg}\n"
        def update_log():
            if not self._is_running:
                return
            try:
                self.log_text.configure(state=tk.NORMAL)
                self.log_text.insert(tk.END, formatted_msg)
                self.log_text.see(tk.END)
                self.log_text.configure(state=tk.DISABLED)
                self.update_detailed_stats()
            except Exception:
                pass
        try:
            self.safe_after(0, update_log)
        except Exception:
            pass

    def update_detailed_stats(self):
        try:
            success_count = 0
            error_count = 0
            skip_count = 0
            
            log_content = self.log_text.get('1.0', tk.END)
            lines = [line.strip() for line in log_content.split('\n') if line.strip()]
            
            for line in lines:
                if '✅' in line and 'Đã comment' in line:
                    success_count += 1
                elif '❌' in line and 'Thất bại' in line:
                    error_count += 1
                elif '⏭️' in line and 'Bỏ qua' in line:
                    skip_count += 1
            
            total_attempts = success_count + error_count + skip_count
            
            stats_text = f"✅ Đã cmt thành công: {success_count} | ❌ Đã cmt thất bại: {error_count} | ⏭️ Đã bỏ qua: {skip_count} | 📊 Tổng số lượt cmt: {total_attempts}"
            
            if total_attempts > 0:
                success_rate = (success_count / total_attempts) * 100
                stats_text += f" | 📈 Tỷ lệ thành công: {success_rate:.1f}%"
            
            self.stats_info_label.configure(text=stats_text)
            
        except Exception as e:
            print(f"Lỗi khi cập nhật thống kê chi tiết: {e}")

    def update_token_log(self, token, message):
        if self.token_manager and self._is_running:
            def do_update():
                try:
                    self.token_manager.update_token_log(token, message)
                except Exception:
                    pass
            try:
                self.safe_after(0, do_update)
            except Exception:
                pass

    def insert_placeholder(self, placeholder):
        current = self.message_var.get()
        if current:
            self.message_var.set(current + " " + placeholder)
        else:
            self.message_var.set(placeholder)

    def on_fetch(self):
        tokens_data = self.gather_tokens_data()
        if not tokens_data:
            try:
                messagebox.showwarning("Thiếu Token", "Vui lòng nhập ít nhất 1 token để tải bài viết!")
            except:
                pass
            return
        uid_inputs = self.uid_source_text.get('1.0', tk.END).strip()
        if not uid_inputs:
            try:
                messagebox.showwarning("Thiếu UID Hoặc Link", "Vui lòng nhập ít nhất một UID hoặc link!")
            except:
                pass
            return
        
        min_reactions = max(0, int(self.min_reactions_var.get()))
        token_info = tokens_data[0]
        token = token_info['token']
        proxy = token_info.get('proxy', '')
        
        uids = []
        for line in uid_inputs.split('\n'):
            line = line.strip()
            if line:
                if "facebook.com" in line:
                    self.log(f"Đang chuyển đổi link sang UID: {line}")
                    try:
                        api_url = f"https://nqtam.id.vn/get-id?link={line}"
                        res = requests.get(api_url, timeout=30).json()
                        if res.get("status") == "success" and res.get("data", {}).get("id"):
                            uid = res["data"]["id"]
                            uids.append(uid)
                            self.log(f"Chuyển đổi thành công → UID: {uid}")
                        else:
                            self.log(f"Lỗi khi chuyển link sang UID: {line}")
                    except Exception:
                        self.log(f"Lỗi khi chuyển link {line}")
                else:
                    uids.append(line)
        
        if not uids:
            self.log("Không có UID hợp lệ!")
            return
        
        limit = max(1, int(self.limit_var.get()))
        
        if min_reactions > 0:
            self.log(f"Đang tải {limit} bài viết từ {len(uids)} UID với reaction tối thiểu: {min_reactions}" + (f" | Proxy: {proxy}" if proxy else ""))
        else:
            self.log(f"Đang tải {limit} bài viết từ {len(uids)} UID..." + (f" | Proxy: {proxy}" if proxy else ""))
        
        def run():
            try:
                try:
                    for w in self.posts_inner.winfo_children():
                        w.destroy()
                except Exception:
                    pass
                
                self.uid_frames.clear()
                if not uids:
                    self.log("Không có UID hợp lệ!")
                    return
                
                for uid in uids:
                    self.log(f"Đang tải bài viết từ UID: {uid}")
                    all_ids = get_post_ids(uid, token, limit * 2, proxy)
                    
                    if not all_ids:
                        self.log(f"Không tìm thấy bài viết nào từ UID: {uid}")
                        continue
                    
                    filtered_ids = []
                    if min_reactions > 0:
                        self.log(f"Đang kiểm tra reaction cho {len(all_ids)} bài viết từ UID: {uid}...")
                        for pid in all_ids:
                            try:
                                reaction_count = get_post_reactions(pid, token, proxy)
                                if reaction_count >= min_reactions:
                                    filtered_ids.append(pid)
                                    self.log(f"✅ {pid} có {reaction_count} reaction - Đủ điều kiện")
                                else:
                                    self.log(f"⏭️ {pid} có {reaction_count} reaction - Không đủ điều kiện")
                                
                                if len(filtered_ids) >= limit:
                                    break
                                    
                            except Exception as e:
                                self.log(f"⚠️ Không thể kiểm tra reaction của {pid}: {e}")
                                filtered_ids.append(pid)
                                if len(filtered_ids) >= limit:
                                    break
                    else:
                        filtered_ids = all_ids[:limit]
                    
                    if not filtered_ids:
                        self.log(f"Không có bài viết nào từ UID: {uid} đủ điều kiện reaction")
                        continue
                    
                    def add_uid_frame(uid=uid, ids=filtered_ids):
                        try:
                            uid_frame = tb.LabelFrame(self.posts_inner, text=f"UID: {uid} - {len(ids)} bài viết" + (f" (reaction ≥ {min_reactions})" if min_reactions > 0 else ""))
                            uid_frame.pack(fill=X, padx=6, pady=6, expand=True)
                            posts_frame = tb.Frame(uid_frame)
                            posts_frame.pack(fill=X, padx=6, pady=6)
                            post_vars = []
                            for pid in ids:
                                var = tk.BooleanVar(value=False)
                                cb = tb.Checkbutton(posts_frame, text=pid, variable=var)
                                cb.pack(anchor="w", padx=6, pady=2)
                                post_vars.append((pid, var))
                            self.uid_frames[uid] = {'frame': uid_frame, 'posts_frame': posts_frame, 'post_vars': post_vars}
                        except Exception:
                            pass
                    
                    try:
                        self.safe_after(0, add_uid_frame)
                    except Exception:
                        pass
                    
                    self.log(f"Đã tải {len(filtered_ids)} bài viết từ UID: {uid}" + (f" (reaction ≥ {min_reactions})" if min_reactions > 0 else ""))
                
                total_posts = sum(len(info.get('post_vars', [])) for info in self.uid_frames.values())
                self.log(f"Đã tải tổng cộng {total_posts} bài viết từ {len(self.uid_frames)} UID" + (f" (reaction ≥ {min_reactions})" if min_reactions > 0 else ""))
                
            except Exception as e:
                self.log(f"Lỗi khi tải bài viết: {e}")
        
        t = threading.Thread(target=run, daemon=True)
        self._threads.append(t)
        t.start()

    def select_all(self):
        for uid_info in self.uid_frames.values():
            for pid, var in uid_info['post_vars']:
                var.set(True)

    def deselect_all(self):
        for uid_info in self.uid_frames.values():
            for pid, var in uid_info['post_vars']:
                var.set(False)

    def gather_tokens_data(self):
        return self.token_manager.get_selected_tokens_data()

    def on_start(self):
        selected = []
        for uid_info in self.uid_frames.values():
            for pid, var in uid_info['post_vars']:
                if var.get():
                    selected.append(pid)
        if not selected:
            try:
                messagebox.showwarning("Chưa Chọn ID Bài Viết", "Chọn ít nhất 1 bài viết để chạy!")
            except:
                pass
            return
        msgs_raw = self.message_var.get().strip()
        if not msgs_raw:
            try:
                messagebox.showwarning("Thiếu Nội Dung Cmt", "Vui lòng nhập nội dung comment!")
            except:
                pass
            return
        messages = [m.strip() for m in msgs_raw.split('|') if m.strip()]
        images = [x.strip() for x in self.images_text.get('1.0', tk.END).splitlines() if x.strip()]
        tokens_data = self.gather_tokens_data()
        if not tokens_data:
            try:
                messagebox.showwarning("Thiếu Token", "Vui lòng thêm ít nhất 1 token!")
            except:
                pass
            return
        for token_info in tokens_data:
            self.update_token_log(token_info['token'], "🟡 Đang chờ..." + (f" | Proxy: {token_info.get('proxy', '')}" if token_info.get('proxy') else ""))
        repeats_per_post = max(1, int(self.repeats_var.get()))
        delay = max(0.0, float(self.delay_var.get()))
        threads_count = max(1, int(self.threads_var.get()))
        token_limit = max(0, int(self.token_limit_var.get()))
        min_reactions = max(0, int(self.min_reactions_var.get()))
        daily_limit = max(0, int(self.daily_limit_var.get()))
        
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(["Token", "Nội dung", "Post ID", "Link Comment", "Proxy"])
        self.ws_rows.clear()
        self.stop_event.clear()
        try:
            self.start_btn.configure(state=DISABLED)
            self.stop_btn.configure(state=NORMAL)
            self.status_label.configure(text="🟡 ĐANG CHẠY...", bootstyle="warning")
        except Exception:
            pass
        token_count = len(tokens_data)
        proxy_count = sum(1 for t in tokens_data if t.get('proxy'))
        post_comment_count = defaultdict(int)
        token_comment_count = defaultdict(int)
        daily_comment_count = defaultdict(int)
        current_date = [datetime.datetime.now().date()]
        
        reaction_info = f" | Reaction tối thiểu: {min_reactions}" if min_reactions > 0 else ""
        daily_info = f" | Giới hạn hàng ngày: {daily_limit}" if daily_limit > 0 else ""
        self.log(f"Bắt đầu: {len(selected)} post | {token_count} token | {proxy_count} proxy | {threads_count} luồng | {repeats_per_post} lần/post | Giới hạn: {token_limit} cmt/token{reaction_info}{daily_info}")
        
        if self.run_mode.get() == "sequential":
            self.log("Chế độ: TUẦN TỰ - Chạy từng token một")
            t = threading.Thread(target=worker_thread_sequential, 
                               args=(selected, tokens_data, messages, images, repeats_per_post, 
                                   delay, self.ws_rows, self.lock, self.log, self.stop_event, 
                                   self.update_token_log, post_comment_count, token_comment_count, token_limit, min_reactions, daily_limit, daily_comment_count, current_date),
                               daemon=True)
            self._threads.append(t)
            t.start()
        else:
            self.log("Chế độ: ĐỒNG LOẠT - Chạy tất cả token cùng lúc")
            chunk_size = len(selected) // threads_count
            if len(selected) % threads_count != 0:
                chunk_size += 1
            chunks = [selected[i:i+chunk_size] for i in range(0, len(selected), chunk_size)]
            for chunk in chunks:
                t = threading.Thread(target=worker_thread_parallel, 
                                   args=(chunk, tokens_data, messages, images, repeats_per_post, 
                                       delay, self.ws_rows, self.lock, self.log, self.stop_event, 
                                       self.update_token_log, post_comment_count, token_comment_count, token_limit, min_reactions, daily_limit, daily_comment_count, current_date),
                                   daemon=True)
                self._threads.append(t)
                t.start()
        t = threading.Thread(target=self.monitor_and_save, daemon=True)
        self._threads.append(t)
        t.start()

    def monitor_and_save(self):
        while True:
            alive = False
            for t in threading.enumerate():
                if t is threading.current_thread():
                    continue
                if t.daemon and t.is_alive():
                    alive = True
                    break
            if not alive or self.stop_event.is_set():
                break
            time.sleep(1)
        def finalize_ui():
            try:
                if self.stop_event.is_set():
                    tokens_data = self.gather_tokens_data()
                    for token_info in tokens_data:
                        self.update_token_log(token_info['token'], "Đã dừng")
                    self.log("Đã dừng!")
                    try:
                        self.status_label.configure(text="🔴 ĐÃ DỪNG", bootstyle="danger")
                    except Exception:
                        pass
                else:
                    self.log("Hoàn thành!")
                    try:
                        self.status_label.configure(text="🟢 HOÀN THÀNH", bootstyle="success")
                    except Exception:
                        pass
                with self.lock:
                    if self.ws_rows: 
                        fname = f"ketqua_cmt_{int(time.time())}.xlsx"
                        try:
                            self.save_links_to_excel(fname)
                            self.log(f"Đã lưu kết quả: {fname}")
                        except Exception as e:
                            self.log(f"Lỗi lưu file: {e}")
                try:
                    self.start_btn.configure(state=NORMAL)
                    self.stop_btn.configure(state=DISABLED)
                except Exception:
                    pass
            except Exception:
                pass
        try:
            self.safe_after(0, finalize_ui)
        except Exception:
            finalize_ui()

    def on_stop(self):
        self.stop_event.set()
        self.log("Đang dừng...")
        tokens_data = self.gather_tokens_data()
        for token_info in tokens_data:
            self.update_token_log(token_info['token'], "Đang dừng...")

    def on_save_excel(self):
        with self.lock:
            if not self.ws_rows:
                try:
                    messagebox.showinfo("Không có dữ liệu", "Chưa có kết quả để lưu!")
                except:
                    pass
                return
        p = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files','*.xlsx')], initialfile=f'ketqua_cmt_{int(time.time())}.xlsx')
        if p:
            with self.lock:
                try:
                    self.save_links_to_excel(p)
                    try:
                        messagebox.showinfo("Lưu thành công", f"Đã lưu file: {p}")
                    except:
                        pass
                except Exception as e:
                    try:
                        messagebox.showerror("Lỗi", f"Lỗi khi lưu file: {e}")
                    except:
                        pass

if __name__ == '__main__':
    root = tb.Window()
    app = FBCommentGUI(root)
    try:
        root.mainloop()
    except KeyboardInterrupt:
        app.on_closing()
    except Exception:
        try:
            app.on_closing()
        except:

            pass
